import csv
import io
import re
import zipfile
from datetime import date
from typing import Any, Dict, List, Literal, Optional
from xml.etree import ElementTree as ET

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from pypdf import PdfReader
from sqlmodel import select, func

from app.database import get_session
from app.models import User
from app.models.attendance_record import AttendanceRecord
from app.auth.deps import get_current_user, role_required

router = APIRouter()

UPLOAD_THRESHOLD = 75.0
EXPORT_HEADERS = [
    "Roll No.",
    "Name of the Student",
    "Total Attended",
    "Out Of",
    "Avg. Attend. %",
    "Status",
]
WORD_NAMESPACE = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


def _normalize_column_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").strip().lower())


def _coerce_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip().replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _coerce_int(value: Any) -> Optional[int]:
    number = _coerce_float(value)
    if number is None:
        return None
    return int(round(number))


def _match_first(row: Dict[str, Any], aliases: List[str]) -> Any:
    normalized = {_normalize_column_name(k): v for k, v in row.items() if k is not None}
    for alias in aliases:
        if alias in normalized and str(normalized[alias]).strip():
            return normalized[alias]
    return None


def _status_for_attendance(attendance_percentage: float) -> str:
    return "Defaulter" if attendance_percentage < UPLOAD_THRESHOLD else "Eligible"


def _normalize_student_row(
    row: Dict[str, Any],
    *,
    fallback_total_hours: Optional[int] = None,
) -> Optional[Dict[str, Any]]:
    roll_no = _match_first(
        row,
        ["rollno", "rollnumber", "rollno.", "roll", "studentid", "studentrollno"],
    )
    name = _match_first(
        row,
        ["nameofthestudent", "studentname", "student", "name"],
    )
    attendance_percentage = _coerce_float(
        _match_first(
            row,
            [
                "avgattend",
                "avgattend%",
                "avgattendance",
                "attendance%",
                "attendancepercentage",
                "avgattendpercent",
            ],
        )
    )
    total_attended = _coerce_int(
        _match_first(
            row,
            [
                "totalattended",
                "totalpresent",
                "dayspresent",
                "hourspresent",
                "present",
                "total",
            ],
        )
    )
    total_hours = _coerce_int(
        _match_first(
            row,
            [
                "outof",
                "totalhours",
                "totalpossible",
                "daystotal",
                "hours",
                "totaloutof",
            ],
        )
    )

    if total_hours is None:
        total_hours = fallback_total_hours
    if attendance_percentage is None and total_attended is not None and total_hours:
        attendance_percentage = round((total_attended / total_hours) * 100, 2)
    if total_attended is None and attendance_percentage is not None and total_hours:
        total_attended = int(round((attendance_percentage / 100) * total_hours))

    if not roll_no or not name or attendance_percentage is None:
        return None

    return {
        "roll_no": str(roll_no).strip(),
        "student_name": str(name).strip(),
        "total_attended": total_attended,
        "total_hours": total_hours,
        "attendance_percentage": round(attendance_percentage, 2),
        "status": _status_for_attendance(attendance_percentage),
    }


def _normalize_student_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    normalized_rows = [_normalize_student_row(row) for row in rows]
    return [row for row in normalized_rows if row]


def _extract_report_metadata_from_pdf(text: str) -> Dict[str, Any]:
    metadata: Dict[str, Any] = {}
    title_match = re.search(r"Defaulter Attendance Sheet \((.*?)\)", text, re.S)
    academic_year_match = re.search(r"Academic Year:\s*([^\n]+)", text)
    class_match = re.search(r"CLASS/DIV:([^\n]+?)Date of Published", text)
    publish_match = re.search(r"Date of Published:\s*([0-9/]+)", text)
    batch_match = re.search(r"Batch-\s*([A-Za-z0-9-]+)", text)
    subject_match = re.search(
        r"Roll No\. Name of the Student\s+Out of \(Hrs\.\)↓\s*(.*?)\s*Mrs\.",
        text,
        re.S,
    )
    total_hours_match = re.search(r"\b240\b", text)

    if title_match:
        metadata["report_period"] = " ".join(title_match.group(1).split())
    if academic_year_match:
        metadata["academic_year"] = academic_year_match.group(1).strip()
    if class_match:
        metadata["class_division"] = " ".join(class_match.group(1).split())
    if publish_match:
        metadata["published_date"] = publish_match.group(1).strip()
    if batch_match:
        metadata["batch"] = batch_match.group(1).strip()
    if subject_match:
        raw_subjects = " ".join(subject_match.group(1).split())
        raw_subjects = raw_subjects.replace("IOTCCLMini Proj", "IOTCCL Mini Proj")
        metadata["subjects"] = re.findall(r"[A-Z][A-Za-z0-9/+-]*(?:\s+[A-Z][A-Za-z0-9/+-]*)*", raw_subjects)
    if total_hours_match:
        metadata["total_hours"] = 240

    return metadata


def _parse_pdf_report(file_bytes: bytes) -> Dict[str, Any]:
    reader = PdfReader(io.BytesIO(file_bytes))
    pages = [(page.extract_text() or "") for page in reader.pages]
    text = "\n".join(pages)
    metadata = _extract_report_metadata_from_pdf(text)
    total_hours = metadata.get("total_hours")
    students: List[Dict[str, Any]] = []

    for line in text.splitlines():
        compact = " ".join(line.split())
        if not compact or not re.match(r"^\d+", compact):
            continue
        parts = compact.split()
        numeric_start = None
        for index, token in enumerate(parts[1:], start=1):
            if re.fullmatch(r"\d+(?:\.\d+)?", token):
                numeric_start = index
                break
        if numeric_start is None:
            continue

        name_tokens = parts[1:numeric_start]
        numeric_tokens = parts[numeric_start:]
        if len(numeric_tokens) < 2:
            continue

        attendance_percentage = _coerce_float(numeric_tokens[-1])
        total_attended = _coerce_int(numeric_tokens[-2])
        if not name_tokens or attendance_percentage is None:
            continue

        students.append(
            {
                "roll_no": parts[0],
                "student_name": " ".join(name_tokens).strip(),
                "total_attended": total_attended,
                "total_hours": total_hours,
                "attendance_percentage": round(attendance_percentage, 2),
                "status": _status_for_attendance(attendance_percentage),
            }
        )

    return {
        "meta": metadata,
        "rows": students,
        "columns": EXPORT_HEADERS,
    }


def _parse_csv_report(file_bytes: bytes) -> Dict[str, Any]:
    decoded = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            decoded = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise HTTPException(status_code=400, detail="Unable to decode CSV file")

    reader = csv.DictReader(io.StringIO(decoded))
    normalized_rows = _normalize_student_rows(list(reader))
    return {
        "meta": {},
        "rows": normalized_rows,
        "columns": EXPORT_HEADERS,
    }


def _parse_html_or_xml_xls_report(file_bytes: bytes) -> Dict[str, Any]:
    decoded = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "utf-16"):
        try:
            decoded = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise HTTPException(status_code=400, detail="Unable to decode Excel file")

    table_match = re.search(r"<table.*?</table>", decoded, re.I | re.S)
    if table_match:
        rows = []
        row_blocks = re.findall(r"<tr.*?</tr>", table_match.group(0), re.I | re.S)
        parsed_rows: List[List[str]] = []
        for row_block in row_blocks:
            cells = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row_block, re.I | re.S)
            parsed_rows.append([re.sub(r"<[^>]+>", "", cell).strip() for cell in cells])

        if parsed_rows:
            headers = parsed_rows[0]
            for values in parsed_rows[1:]:
                row = {headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers))}
                rows.append(row)
            return {
                "meta": {"sheet_name": "Sheet1"},
                "rows": _normalize_student_rows(rows),
                "columns": EXPORT_HEADERS,
            }

    raise HTTPException(
        status_code=400,
        detail="Unsupported legacy .xls file. Please upload .xlsx, CSV, PDF, or DOCX.",
    )


def _parse_excel_report(file_bytes: bytes) -> Dict[str, Any]:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        return _parse_html_or_xml_xls_report(file_bytes)
    sheet = workbook.active
    raw_rows = list(sheet.iter_rows(values_only=True))
    if not raw_rows:
        raise HTTPException(status_code=400, detail="Excel sheet is empty")

    headers = [str(cell).strip() if cell is not None else "" for cell in raw_rows[0]]
    rows: List[Dict[str, Any]] = []
    for values in raw_rows[1:]:
        row = {headers[idx]: value for idx, value in enumerate(values) if idx < len(headers)}
        normalized = _normalize_student_row(row)
        if normalized:
            rows.append(normalized)

    return {
        "meta": {"sheet_name": sheet.title},
        "rows": rows,
        "columns": EXPORT_HEADERS,
    }


def _parse_docx_report(file_bytes: bytes) -> Dict[str, Any]:
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
            document_xml = archive.read("word/document.xml")
    except (KeyError, zipfile.BadZipFile):
        raise HTTPException(status_code=400, detail="Invalid DOCX file")

    root = ET.fromstring(document_xml)
    rows: List[Dict[str, Any]] = []
    metadata: Dict[str, Any] = {}

    for table in root.findall(".//w:tbl", WORD_NAMESPACE):
        extracted_rows: List[List[str]] = []
        for tr in table.findall("./w:tr", WORD_NAMESPACE):
            values = []
            for tc in tr.findall("./w:tc", WORD_NAMESPACE):
                text = "".join(tc.itertext()).strip()
                values.append(" ".join(text.split()))
            if any(values):
                extracted_rows.append(values)

        if len(extracted_rows) >= 2:
            headers = extracted_rows[0]
            table_rows: List[Dict[str, Any]] = []
            for values in extracted_rows[1:]:
                row = {
                    headers[idx]: values[idx] if idx < len(values) else ""
                    for idx in range(len(headers))
                }
                table_rows.append(row)
            normalized = _normalize_student_rows(table_rows)
            if normalized:
                rows.extend(normalized)
                break

    if not rows:
        paragraph_text = "\n".join(
            " ".join("".join(paragraph.itertext()).split())
            for paragraph in root.findall(".//w:p", WORD_NAMESPACE)
        )
        line_rows: List[Dict[str, Any]] = []
        for line in paragraph_text.splitlines():
            compact = " ".join(line.split())
            if not compact or not re.match(r"^\d+", compact):
                continue
            match = re.match(r"^(\S+)\s+(.+?)\s+(\d+(?:\.\d+)?)\s+(\d+(?:\.\d+)?)\s+(\d+(?:\.\d+)?)$", compact)
            if not match:
                continue
            line_rows.append(
                {
                    "roll no": match.group(1),
                    "name of the student": match.group(2),
                    "total attended": match.group(3),
                    "out of": match.group(4),
                    "avg. attend. %": match.group(5),
                }
            )
        rows = _normalize_student_rows(line_rows)

    if not rows:
        raise HTTPException(
            status_code=400,
            detail="Could not extract attendance rows from DOCX. Use a tabular DOCX, CSV, XLSX, or PDF file.",
        )

    return {
        "meta": metadata,
        "rows": rows,
        "columns": EXPORT_HEADERS,
    }


def _build_upload_response(parsed: Dict[str, Any], filename: str, source_type: str) -> Dict[str, Any]:
    rows = parsed["rows"]
    defaulters = [
        row for row in rows if row["attendance_percentage"] < UPLOAD_THRESHOLD
    ]
    return {
        "source_file": filename,
        "source_type": source_type,
        "threshold": UPLOAD_THRESHOLD,
        "columns": parsed["columns"],
        "meta": parsed.get("meta", {}),
        "rows": rows,
        "defaulters": defaulters,
        "summary": {
            "student_count": len(rows),
            "defaulter_count": len(defaulters),
        },
    }


def _report_rows_to_export(rows: List[Dict[str, Any]]) -> List[List[Any]]:
    return [
        [
            row.get("roll_no", ""),
            row.get("student_name", ""),
            row.get("total_attended", ""),
            row.get("total_hours", ""),
            row.get("attendance_percentage", ""),
            row.get("status", ""),
        ]
        for row in rows
    ]


class AttendanceReportRow(BaseModel):
    roll_no: str
    student_name: str
    total_attended: Optional[int] = None
    total_hours: Optional[int] = None
    attendance_percentage: float
    status: str


class AttendanceReportExportPayload(BaseModel):
    format: Literal["csv", "xlsx", "docx"]
    rows: List[AttendanceReportRow]
    source_file: Optional[str] = None


class AttendanceCreate(BaseModel):
    student_id: int
    class_id: str
    day: date
    present: bool


class AttendanceRead(BaseModel):
    id: int
    student_id: int
    class_id: str
    day: date
    present: bool


class BulkAttendanceUpload(BaseModel):
    class_id: str
    day: date
    records: List[dict]  # [{"student_id": 1, "present": true}, ...]


@router.post(
    "/report/import",
    summary="Import attendance report file",
    description="Upload CSV, XLSX, or PDF attendance report and extract defaulters below 75%",
)
async def import_attendance_report(file: UploadFile = File(...)):
    filename = file.filename or "attendance-report"
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    if extension == "csv":
        parsed = _parse_csv_report(file_bytes)
        source_type = "csv"
    elif extension in {"xlsx", "xls"}:
        parsed = _parse_excel_report(file_bytes)
        source_type = extension
    elif extension == "pdf":
        parsed = _parse_pdf_report(file_bytes)
        source_type = "pdf"
    elif extension == "docx":
        parsed = _parse_docx_report(file_bytes)
        source_type = "docx"
    else:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Upload CSV, XLS, XLSX, PDF, or DOCX.",
        )

    return _build_upload_response(parsed, filename, source_type)


@router.post(
    "/report/export",
    summary="Export attendance report",
    description="Export the uploaded attendance report rows as CSV or Excel",
)
def export_attendance_report(payload: AttendanceReportExportPayload):
    rows = [row.model_dump() for row in payload.rows]
    filename_stem = (payload.source_file or "attendance_report").rsplit(".", 1)[0]

    if payload.format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(EXPORT_HEADERS)
        writer.writerows(_report_rows_to_export(rows))
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{filename_stem}_defaulters.csv"'
            },
        )

    if payload.format == "docx":
        document_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:wpc="http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas"
 xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"
 xmlns:o="urn:schemas-microsoft-com:office:office"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
 xmlns:m="http://schemas.openxmlformats.org/officeDocument/2006/math"
 xmlns:v="urn:schemas-microsoft-com:vml"
 xmlns:wp14="http://schemas.microsoft.com/office/word/2010/wordprocessingDrawing"
 xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"
 xmlns:w10="urn:schemas-microsoft-com:office:word"
 xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"
 xmlns:w14="http://schemas.microsoft.com/office/word/2010/wordml"
 xmlns:wpg="http://schemas.microsoft.com/office/word/2010/wordprocessingGroup"
 xmlns:wpi="http://schemas.microsoft.com/office/word/2010/wordprocessingInk"
 xmlns:wne="http://schemas.microsoft.com/office/word/2006/wordml"
 xmlns:wps="http://schemas.microsoft.com/office/word/2010/wordprocessingShape"
 mc:Ignorable="w14 wp14">
 <w:body>
  <w:p><w:r><w:t>Attendance Defaulters Report</w:t></w:r></w:p>
  <w:tbl>
   {"".join(
            "<w:tr>" +
            "".join(
                f"<w:tc><w:p><w:r><w:t>{value}</w:t></w:r></w:p></w:tc>"
                for value in row_values
            ) +
            "</w:tr>"
            for row_values in [EXPORT_HEADERS, *_report_rows_to_export(rows)]
        )}
  </w:tbl>
  <w:sectPr><w:pgSz w:w=\"12240\" w:h=\"15840\"/><w:pgMar w:top=\"1440\" w:right=\"1440\" w:bottom=\"1440\" w:left=\"1440\"/></w:sectPr>
 </w:body>
</w:document>"""
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr(
                "[Content_Types].xml",
                """<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
 <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
 <Default Extension="xml" ContentType="application/xml"/>
 <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>""",
            )
            archive.writestr(
                "_rels/.rels",
                """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
 <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>""",
            )
            archive.writestr("word/document.xml", document_xml)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={
                "Content-Disposition": f'attachment; filename="{filename_stem}_defaulters.docx"'
            },
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Defaulters"
    sheet.append(EXPORT_HEADERS)
    for row in _report_rows_to_export(rows):
        sheet.append(row)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename_stem}_defaulters.xlsx"'
        },
    )


@router.post(
    "/bulk-upload",
    response_model=List[AttendanceRead],
    status_code=201,
    summary="Bulk upload attendance",
    description="Teacher uploads CSV with student_id,present columns",
)
def bulk_upload_attendance(
    payload: BulkAttendanceUpload,
    current_user: User = Depends(role_required(["teacher", "admin"])),
    session=Depends(get_session),
):
    created_records = []
    for record in payload.records:
        student_id = record.get("student_id")
        present = record.get("present", True)

        if not student_id:
            continue

        student = session.get(User, student_id)
        if not student or student.role != "student":
            continue

        existing = session.exec(
            select(AttendanceRecord).where(
                AttendanceRecord.student_id == student_id,
                AttendanceRecord.class_id == payload.class_id,
                AttendanceRecord.day == payload.day,
            )
        ).first()

        if existing:
            existing.present = present
            session.add(existing)
            created_records.append(existing)
        else:
            new_record = AttendanceRecord(
                student_id=student_id,
                class_id=payload.class_id,
                day=payload.day,
                present=present,
            )
            session.add(new_record)
            created_records.append(new_record)

    session.commit()
    for r in created_records:
        session.refresh(r)

    return [
        AttendanceRead(
            id=r.id,
            student_id=r.student_id,
            class_id=r.class_id,
            day=r.day,
            present=r.present,
        )
        for r in created_records
    ]


@router.get(
    "/class/{class_id}",
    response_model=List[AttendanceRead],
    summary="Get class attendance",
    description="Get attendance records for a class",
)
def get_class_attendance(
    class_id: str,
    day: Optional[date] = None,
    current_user: User = Depends(role_required(["teacher", "admin"])),
    session=Depends(get_session),
):
    query = select(AttendanceRecord).where(AttendanceRecord.class_id == class_id)
    if day:
        query = query.where(AttendanceRecord.day == day)

    records = session.exec(query.order_by(AttendanceRecord.day.desc())).all()

    return [
        AttendanceRead(
            id=r.id,
            student_id=r.student_id,
            class_id=r.class_id,
            day=r.day,
            present=r.present,
        )
        for r in records
    ]


@router.get(
    "/defaulters/{class_id}",
    response_model=List[dict],
    summary="Get defaulters",
    description="Get students with attendance below 75%",
)
def get_defaulters(
    class_id: str,
    current_user: User = Depends(role_required(["teacher", "admin"])),
    session=Depends(get_session),
):
    # Get all students in class
    students = session.exec(
        select(User).where(User.class_id == class_id, User.role == "student")
    ).all()

    if not students:
        return []

    defaulter_list = []
    for student in students:
        # Get total days attended
        total_days = (
            session.exec(
                select(func.count(AttendanceRecord.id)).where(
                    AttendanceRecord.student_id == student.id,
                    AttendanceRecord.class_id == class_id,
                )
            ).one()
            or 0
        )

        if total_days == 0:
            continue

        present_days = (
            session.exec(
                select(func.count(AttendanceRecord.id)).where(
                    AttendanceRecord.student_id == student.id,
                    AttendanceRecord.class_id == class_id,
                    AttendanceRecord.present == True,
                )
            ).one()
            or 0
        )

        attendance_pct = (present_days / total_days) * 100 if total_days > 0 else 0

        if attendance_pct < 75:
            defaulter_list.append(
                {
                    "student_id": student.id,
                    "student_name": student.name,
                    "roll_no": getattr(student, "roll_no", None) or student.email,
                    "present_days": present_days,
                    "total_days": total_days,
                    "attendance_percentage": round(attendance_pct, 2),
                }
            )

    return defaulter_list


@router.get(
    "/export/{class_id}",
    summary="Export attendance CSV",
    description="Export attendance as CSV",
)
def export_attendance(
    class_id: str,
    current_user: User = Depends(role_required(["teacher", "admin"])),
    session=Depends(get_session),
):
    records = session.exec(
        select(AttendanceRecord)
        .where(AttendanceRecord.class_id == class_id)
        .order_by(AttendanceRecord.day)
    ).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["student_id", "class_id", "day", "present"])

    for r in records:
        writer.writerow([r.student_id, r.class_id, r.day, r.present])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=attendance_{class_id}.csv"
        },
    )
